import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def extrair_dados(pdf_file, buscar, excel_file_path, nome_unidade):
   
    # Inicializar dicionário para armazenar dados extraídos com valores 0
    dados_extraidos = {palavra_ou_frase: 0 for palavra_ou_frase in buscar}
    
    grandezas = ["GB", "TB", "MB", "KB"]

    # Abrir o arquivo PDF
    with pdfplumber.open(pdf_file) as pdf:
        if len(pdf.pages) >= 2:
            # Extrair texto da segunda página
            segunda_pagina = pdf.pages[1]
            texto = segunda_pagina.extract_text()
            linhas = texto.split('\n')

            # Processar cada linha
            for linha in linhas:
                for palavra_ou_frase in buscar:
                    if palavra_ou_frase in linha:
                        palavras = linha.split()
                        encontrou_numero = False

                        for i, palavra in enumerate(palavras):
                            try:
                                # Tentar converter a palavra para float
                                numero = float(palavra)
                                
                                if i > 0:
                                    texto_antes_do_numero = ' '.join(palavras[i - len(palavra_ou_frase.split()):i])
                                    if texto_antes_do_numero == palavra_ou_frase:
                                        # Verificar a grandeza após o número
                                        grandeza = ""
                                        if i + 1 < len(palavras):
                                            proxima_palavra = palavras[i + 1]
                                            if proxima_palavra in grandezas:
                                                grandeza = proxima_palavra
                                        
                                        # Armazenar o número e a grandeza como dados extraídos
                                        dados_extraidos[palavra_ou_frase] = f"{numero} {grandeza}"
                                        encontrou_numero = True
                                        break
                            except ValueError:
                                continue

    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook["Geral"]

    # Encontrar a próxima linha vazia na planilha
    proxima_linha = sheet.max_row + 1

    # Adicionar a unidade correspondente na coluna "Unidade"
    for coluna in range(1, sheet.max_column + 1):
        coluna_letra = get_column_letter(coluna)
        celula = sheet[f"{coluna_letra}1"]
        
        if celula.value == "Unidade":
            sheet[f"{coluna_letra}{proxima_linha}"].value = nome_unidade
            break

    # Adicionar a data atual na coluna "Data"
    data_atual = datetime.today().strftime('%d/%m/%Y')
    for coluna in range(1, sheet.max_column + 1):
        coluna_letra = get_column_letter(coluna)
        celula = sheet[f"{coluna_letra}1"]
        
        if celula.value == "Data":
            sheet[f"{coluna_letra}{proxima_linha}"].value = data_atual
            break

    # Adicionar os dados extraídos à planilha "Geral"
    for palavra_ou_frase in buscar:
        for coluna in range(1, sheet.max_column + 1):
            coluna_letra = get_column_letter(coluna)
            celula = sheet[f"{coluna_letra}1"]
            
            if celula.value == palavra_ou_frase:
                # Se os dados extraídos são None, substitua por zero
                valor = dados_extraidos[palavra_ou_frase]
                if valor is None:
                    valor = 0
                nova_celula = sheet[f"{coluna_letra}{proxima_linha}"]
                nova_celula.value = valor
                break

    # Salvar o arquivo Excel
    workbook.save(excel_file_path)

# Lista de caminhos para os arquivos PDF
pdf_paths = [
    './GRU/abril/Firewall GRU - Report.pdf',
    './BA/abril/Firewall Bdo - Report.pdf',
    './SP/abril/Firewall SP - Report.pdf'
]

# Lista de nomes das unidades correspondentes
unidades = [
    'Guarulhos',
    'Brumado',
    'São Paulo'
]

# Caminho para o arquivo Excel
excel_file_path = './test.xlsx'

# Lista de palavras ou frases para buscar nos PDFs
buscar = [
    'Web domains accessed',
    'Web domains blocked',
    'Packets blocked by Firewall',
    'Attacks blocked by IPS',
    'Total Website requests',
    'URLs blocked',
    'Uplink fail-overs',
    'HTTP/S Malware blocked',
    'Applications accessed',
    'Blocked Applications',
    'Intrusion Attacks',
    'Emergency + Critical Attacks',
    'Application Data Transfer',
    'Web data transfer',
    'Total User Data Transfer',
    'App Risk Score (out of 5)'
]

# Extrair dados de cada PDF
for i, pdf_path in enumerate(pdf_paths):
    extrair_dados(pdf_path, buscar, excel_file_path, unidades[i])

print("Dados extraidos com sucesso!")
